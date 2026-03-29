"""Excel workbook creation, formatting, and update logic."""

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Color, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .paths import get_archive_dir, resolve_relative

logger = logging.getLogger(__name__)

# ── Styles ────────────────────────────────────────────────────────────────────

_BLUE_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10)
_YELLOW_FILL = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
_TUESDAY_FILL = PatternFill(patternType="solid", fgColor=Color(theme=9, tint=0.3999755851924192))
_GREEN_FONT = Font(color="006100")
_RED_FONT = Font(color="9C0006")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Column spec: (header_text, width, header_group)
# header_group: 'blue', 'yellow', 'plain'
COLUMNS = [
    ("Day",                          13.0,  "plain"),   # A
    ("Date",                         11.0,  "blue"),    # B
    ("Open",                         10.0,  "blue"),    # C
    ("High",                         13.0,  "blue"),    # D
    ("Low",                          13.0,  "blue"),    # E
    ("Close",                        13.0,  "blue"),    # F
    ("Daily Change",                  9.5,  "blue"),    # G
    ("Weekly Change",                13.0,  "blue"),    # H
    ("Close Above\nprev High",      12.0,  "yellow"),  # I
    ("Close Below\nprev Low",       12.0,  "yellow"),  # J
    ("HH / HL",                     10.5,  "yellow"),  # K
    ("LL / LH",                     10.5,  "yellow"),  # L
    ("Diff in\nFuture",              9.5,  "plain"),   # M
    ("Future\nas on Date",          13.0,  "plain"),   # N
    ("Future\nExpiry",              13.0,  "plain"),   # O
    ("Comments",                    30.0,  "plain"),   # P — user notes (never overwritten)
]


def _apply_header_style(ws) -> None:
    """Apply header styles and column widths."""
    for col_idx, (text, width, group) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

        if group == "blue":
            cell.fill = _BLUE_FILL
            cell.font = _WHITE_BOLD
            cell.alignment = _CENTER_WRAP
        elif group == "yellow":
            cell.fill = _YELLOW_FILL
            cell.font = _WHITE_BOLD
            cell.alignment = _CENTER_WRAP
        else:
            cell.font = Font(bold=True, size=10)
            cell.alignment = _CENTER_WRAP


def create_workbook(excel_path: str) -> None:
    """Create a new, empty Excel workbook with formatted headers."""
    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "NSE_Data"

    _apply_header_style(ws)

    # Freeze the top row so headers stay visible when scrolling
    ws.freeze_panes = "A2"

    wb.save(str(path))
    logger.info("Created new workbook: %s", excel_path)


def backup_workbook(excel_path: str) -> Optional[str]:
    """Create a timestamped backup copy in archive/. Returns backup path or None."""
    src = Path(excel_path)
    if not src.exists():
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = get_archive_dir() / f"{src.stem}_{ts}{src.suffix}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(src), str(dest))
    logger.info("Backup saved: %s", dest)
    return str(dest)


def get_last_date(excel_path: str) -> Optional[datetime]:
    """Read the last date entry from column B of the workbook."""
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        last_row = ws.max_row
        if last_row <= 1:
            wb.close()
            return None
        val = ws.cell(last_row, 2).value  # Column B = Date
        wb.close()
        if isinstance(val, datetime):
            return val
        if isinstance(val, str):
            return pd.to_datetime(val)
        return None
    except Exception as exc:
        logger.error("Error reading last date: %s", exc)
        return None


def _apply_all_tuesday_highlights(ws) -> None:
    """Scan ALL data rows and highlight entire Tuesday rows (all columns A-P).

    Only applies fill to Tuesday rows — non-Tuesday rows are left untouched
    so that user-applied formatting and cell comments are preserved.
    """
    num_cols = len(COLUMNS)  # 16 (A through P)
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val is None:
            continue
        try:
            if isinstance(val, datetime):
                dt = val
            elif isinstance(val, str):
                dt = datetime.strptime(val[:10], "%Y-%m-%d")
            else:
                continue
        except (ValueError, TypeError):
            continue

        if dt.weekday() == 1:  # Tuesday
            for col in range(1, num_cols + 1):
                ws.cell(row, col).fill = _TUESDAY_FILL


def _add_conditional_formatting(ws) -> None:
    """Add sheet-level conditional formatting rules for change columns.

    Rules skip Tuesday rows (WEEKDAY($B2,2)=2) so the Tuesday highlight
    fill is never overridden by the green/red formatting.
    """
    from openpyxl.formatting.rule import FormulaRule

    # Clear existing rules to prevent accumulation across repeated runs
    ws.conditional_formatting._cf_rules.clear()

    # Daily Change (G) — green if > 0, red if < 0 (skip Tuesdays)
    ws.conditional_formatting.add(
        "G2:G1048576",
        FormulaRule(formula=['AND(G2>0,WEEKDAY($B2,2)<>2)'], fill=_GREEN_FILL, font=_GREEN_FONT),
    )
    ws.conditional_formatting.add(
        "G2:G1048576",
        FormulaRule(formula=['AND(G2<0,WEEKDAY($B2,2)<>2)'], fill=_RED_FILL, font=_RED_FONT),
    )
    # Weekly Change (H)
    ws.conditional_formatting.add(
        "H2:H1048576",
        FormulaRule(formula=['AND(H2>0,WEEKDAY($B2,2)<>2)'], fill=_GREEN_FILL, font=_GREEN_FONT),
    )
    ws.conditional_formatting.add(
        "H2:H1048576",
        FormulaRule(formula=['AND(H2<0,WEEKDAY($B2,2)<>2)'], fill=_RED_FILL, font=_RED_FONT),
    )
    # Difference in future (M)
    ws.conditional_formatting.add(
        "M2:M1048576",
        FormulaRule(formula=['AND(M2>0,WEEKDAY($B2,2)<>2)'], fill=_GREEN_FILL, font=_GREEN_FONT),
    )
    ws.conditional_formatting.add(
        "M2:M1048576",
        FormulaRule(formula=['AND(M2<0,WEEKDAY($B2,2)<>2)'], fill=_RED_FILL, font=_RED_FONT),
    )


def update_workbook(
    excel_path: str,
    data: pd.DataFrame,
    futures_price: Optional[float] = None,
    futures_expiry: Optional[str] = None,
) -> tuple[int, int]:
    """Append new OHLC rows to the workbook, skipping duplicates.

    Futures data is written to the most recent row (last added, or last
    existing row if no new data was added). This ensures futures are
    recorded even when running on weekends/holidays.

    Returns (added_count, skipped_count).
    """
    try:
        wb = load_workbook(excel_path)
    except PermissionError:
        msg = (f"Cannot open '{excel_path}' — the file is locked. "
               "Please close it in Excel and try again.")
        logger.error(msg)
        raise PermissionError(msg)
    ws = wb.active

    # Build date-to-row map for duplicate detection and weekly change lookup
    date_row_map: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val:
            if isinstance(val, datetime):
                date_row_map[val.strftime("%Y-%m-%d")] = row
            elif isinstance(val, str):
                date_row_map[str(val)[:10]] = row

    last_row = ws.max_row
    added = 0
    skipped = 0
    latest_added_row: Optional[int] = None

    for date_idx, row_data in data.iterrows():
        date_str = date_idx.strftime("%Y-%m-%d")

        if date_str in date_row_map:
            logger.debug("Skipping duplicate: %s", date_str)
            skipped += 1
            continue

        last_row += 1
        added += 1
        latest_added_row = last_row

        # A: Day formula (works because B stores a real datetime)
        ws.cell(last_row, 1, f'=TEXT(B{last_row},"dddd")')
        ws.cell(last_row, 1).alignment = _CENTER

        # B: Date — store as actual datetime so TEXT() and number_format work
        ws.cell(last_row, 2, date_idx.to_pydatetime().replace(tzinfo=None))
        ws.cell(last_row, 2).number_format = "YYYY-MM-DD"
        ws.cell(last_row, 2).alignment = _CENTER

        # C–F: OHLC
        for col_off, field in enumerate(["Open", "High", "Low", "Close"]):
            cell = ws.cell(last_row, 3 + col_off, round(row_data[field], 2))
            cell.number_format = "#,##0.00"
            cell.alignment = _CENTER

        # G: Daily Change = Close - prev Close (skip first data row to avoid #VALUE!)
        if last_row > 2:
            ws.cell(last_row, 7, f"=F{last_row}-F{last_row - 1}")
            ws.cell(last_row, 7).number_format = "#,##0.00"
            ws.cell(last_row, 7).alignment = _CENTER

        # H: Weekly Change — exactly 5 trading days back (holiday-aware)
        prior_dates = sorted(d for d in date_row_map if d < date_str)
        if len(prior_dates) >= 5:
            ref_date = prior_dates[-5]
            ref_row = date_row_map[ref_date]
            ws.cell(last_row, 8, f"=F{last_row}-F{ref_row}")
            ws.cell(last_row, 8).number_format = "#,##0.00"
            ws.cell(last_row, 8).alignment = _CENTER

        # Track this row in date_row_map for subsequent weekly-change lookups
        date_row_map[date_str] = last_row

        # I–L: Technical analysis (needs previous row)
        if last_row > 2:
            prev = last_row - 1
            # I: Close above previous High
            ws.cell(last_row, 9, f'=IF(F{last_row}>D{prev},"High","-")')
            ws.cell(last_row, 9).alignment = _CENTER
            # J: Close below previous Low
            ws.cell(last_row, 10, f'=IF(F{last_row}<E{prev},"Low","-")')
            ws.cell(last_row, 10).alignment = _CENTER
            # K: HH / HL
            ws.cell(last_row, 11, f'=IF(D{last_row}>D{prev},"Higher High","LH")')
            ws.cell(last_row, 11).alignment = _CENTER
            # L: LL / LH
            ws.cell(last_row, 12, f'=IF(E{last_row}<E{prev},"L Low","Higher Low")')
            ws.cell(last_row, 12).alignment = _CENTER

        # Thin bottom border for readability (A–O only, skip P to preserve user comments)
        for col in range(1, 16):
            ws.cell(last_row, col).border = _THIN_BORDER

    # M–O: Futures — write on the most recent row (last added or last existing)
    if futures_price is not None:
        fut_row = latest_added_row or ws.max_row
        if fut_row > 1:  # Skip if only header
            ws.cell(fut_row, 14, futures_price)
            ws.cell(fut_row, 14).number_format = "#,##0.00"
            ws.cell(fut_row, 14).alignment = _CENTER

            if futures_expiry:
                ws.cell(fut_row, 15, futures_expiry)
                ws.cell(fut_row, 15).alignment = _CENTER

            # M: Difference = Future - Close
            ws.cell(fut_row, 13, f"=N{fut_row}-F{fut_row}")
            ws.cell(fut_row, 13).number_format = "#,##0.00"
            ws.cell(fut_row, 13).alignment = _CENTER
            logger.debug("Futures written to row %d.", fut_row)

    # Apply Tuesday highlighting to ALL rows (handles existing + new data)
    _apply_all_tuesday_highlights(ws)

    # Apply conditional formatting rules (clears old rules first)
    _add_conditional_formatting(ws)

    # Ensure freeze panes
    ws.freeze_panes = "A2"

    try:
        wb.save(excel_path)
    except PermissionError:
        msg = (f"Cannot save '{excel_path}' — the file is locked. "
               "Please close it in Excel and try again.")
        logger.error(msg)
        wb.close()
        raise PermissionError(msg)
    wb.close()

    if added:
        logger.info("Added %d new row(s) to %s.", added, excel_path)
    if skipped:
        logger.info("Skipped %d duplicate row(s).", skipped)
    if not added and not skipped:
        logger.info("No new rows to add.")

    return added, skipped
